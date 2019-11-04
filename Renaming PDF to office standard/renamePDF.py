import os
import sys
import re
from openpyxl import load_workbook
from PyQt4 import QtCore, QtGui

try:
    _fromUtf8 = QtCore.QString.fromUtf8
except AttributeError:
    def _fromUtf8(s):
        return s

try:
    _encoding = QtGui.QApplication.UnicodeUTF8
    def _translate(context, text, disambig):
        return QtGui.QApplication.translate(context, text, disambig, _encoding)
except AttributeError:
    def _translate(context, text, disambig):
        return QtGui.QApplication.translate(context, text, disambig)

class Ui_Form(QtGui.QMainWindow):

    def __init__(self):
        super(Ui_Form, self).__init__()
        self.setupUi

    def setupUi(self, Form):
        Form.setObjectName(_fromUtf8("Form"))
        Form.resize(233, 167)
        self.verticalLayout_2 = QtGui.QVBoxLayout(Form)
        self.verticalLayout_2.setObjectName(_fromUtf8("verticalLayout_2"))
        self.groupBox = QtGui.QGroupBox(Form)
        font = QtGui.QFont()
        font.setFamily(_fromUtf8("Arial"))
        font.setPointSize(10)
        self.groupBox.setFont(font)
        self.groupBox.setObjectName(_fromUtf8("groupBox"))
        self.verticalLayoutWidget = QtGui.QWidget(self.groupBox)
        self.verticalLayoutWidget.setGeometry(QtCore.QRect(10, 30, 191, 101))
        self.verticalLayoutWidget.setObjectName(_fromUtf8("verticalLayoutWidget"))
        self.verticalLayout = QtGui.QVBoxLayout(self.verticalLayoutWidget)
        self.verticalLayout.setObjectName(_fromUtf8("verticalLayout"))

        self.fromRevit = QtGui.QPushButton(self.verticalLayoutWidget)
        self.fromRevit.setObjectName(_fromUtf8("fromRevit"))
        self.fromRevit.clicked.connect(lambda: self.fromRevitPDF())


        self.verticalLayout.addWidget(self.fromRevit)

        self.fromPDF = QtGui.QPushButton(self.verticalLayoutWidget)
        self.fromPDF.setObjectName(_fromUtf8("fromPDF"))
        self.fromPDF.clicked.connect(lambda: self.fromExisting())

        self.verticalLayout.addWidget(self.fromPDF)
        self.verticalLayout_2.addWidget(self.groupBox)

        self.retranslateUi(Form)
        QtCore.QMetaObject.connectSlotsByName(Form)

    def retranslateUi(self, Form):
        Form.setWindowTitle(_translate("Form", "Form", None))
        self.groupBox.setTitle(_translate("Form", "Rename PDF", None))
        self.fromRevit.setText(_translate("Form", "From Revit", None))
        self.fromPDF.setText(_translate("Form", "Existing Sheet", None))

    def fromRevitPDF(self):
        self.name = QtGui.QFileDialog.getOpenFileNames(self, "Select Files")
        locDir = os.path.dirname(str(self.name[0]))
        os.chdir(locDir)

        for f in self.name:
            f_name, f_ext = os.path.splitext(f)
            if f_ext == ".pdf":
                f_split = f_name.split()
                for f_namsplit in f_split:
                    wb = load_workbook('C:\Revit data\Raw data.xlsx')
                    for i in range(len(wb.get_sheet_names())):
                        wb.active = i
                        if wb.active.title == "Sheets":
                            ws = wb.active
                            for row in ws.iter_rows():
                                sheets = row[0].value
                                if sheets == f_namsplit:
                                    revisionNum = row[2].value
                                    new_name = "{}{}{}{}{}".format(sheets, " (", revisionNum, ")", f_ext)
                                    print(new_name)

                                    os.rename(f, new_name)


        sys.exit()

    def fromExisting(self):
        self.name = QtGui.QFileDialog.getOpenFileNames(self, "Select Files")
        locDir = os.path.dirname(str(self.name[0]))
        os.chdir(locDir)

        for f in self.name:
            f_name, f_ext = os.path.splitext(f)
            if f_ext == ".pdf":
                f_name_1, f_ext_1 = os.path.splitext(f_name)
                f_extsheetNum = f_name_1.split("\\")
                f_sheetNum = f_extsheetNum[-1].split("(")[0].strip()
                wb = load_workbook('C:\Revit data\Raw data.xlsx')
                for i in range(len(wb.get_sheet_names())):
                    wb.active = i
                    if wb.active.title == "Sheets":
                        ws = wb.active
                        for row in ws.iter_rows():
                            sheets = row[0].value
                            if sheets == f_sheetNum:
                                revisionNum = row[2].value
                                new_name = "{}{}{}{}{}".format(sheets, " (", revisionNum, ")", f_ext)
                                print(new_name)

                                os.rename(f, new_name)

        sys.exit()

if __name__ == "__main__":
    import sys
    app = QtGui.QApplication(sys.argv)
    Form = QtGui.QWidget()
    ui = Ui_Form()
    ui.setupUi(Form)
    Form.show()
    sys.exit(app.exec_())