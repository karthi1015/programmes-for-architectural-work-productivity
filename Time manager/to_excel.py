import pandas as pd
import sqlite3
import calendar
import csv_to_db

from openpyxl import load_workbook
from datetime import datetime, timedelta, date, time

conn = sqlite3.connect('schedule.db')
c = conn.cursor()

df = pd.read_sql('SELECT * FROM engenium_projects', conn)

def hours_spent_func(data_frame):
    seconds = 0
    for ea_second in df[data_frame]['hours_spent']:
        seconds = seconds + ea_second
    mins, seconds = divmod(seconds, 60)
    hours, mins = divmod(mins, 60)
    return('%02d:%02d:%02d' % (hours, mins, seconds))

def print_time_code(data_frame):
    print('Time code used')
    return(df[data_frame]['time_code'].drop_duplicates().to_string(index=False))



start_date_input = str(input("Week starting date (e.g 2017-7-1) : "))

full_date_input = datetime.strptime(start_date_input, '%Y-%m-%d')
year = int(full_date_input.year)
month = int(full_date_input.month)
day = int(full_date_input.day)

wb = load_workbook('timesheet_template.xlsm')
ws = wb['TIMESHEET']

for day_num in range(0,7):

    d_week = day + day_num
    if d_week == calendar.monthrange(year,month)[1] + 1:
        break
    else:
        date_to_week = datetime(year,month,d_week)
        week_day = date_to_week.weekday()

        # Mon = 0, Tue = 1, Wed = 2, Thu = 3, Fri = 4, Sat = 5, Sun = 6

        each_date = df[df['start_date'] == str(date_to_week.date())]
        filter_each_date = each_date[['project_number','activity','time_code', 'start_date','start_time','end_time','description']]
        for index, row_col in filter_each_date.iterrows():
            row_full_date = row_col['start_date']
            row_date = datetime.strptime(row_full_date,'%Y-%m-%d')
            row_year = int(row_date.year)
            row_month = int(row_date.month)
            row_day = int(row_date.day)
            row_date_to_week = datetime(row_year, row_month, row_day)
            row_week_day = row_date_to_week.weekday()
            # print(row_date_to_week.date(),row_week_day)

            if row_week_day == 0:
                ex_col = 10
                ex_col_time = 11
            elif row_week_day == 1:
                ex_col = 12
                ex_col_time = 13
            elif row_week_day == 2:
                ex_col = 14
                ex_col_time = 15
            elif row_week_day == 3:
                ex_col = 2
                ex_col_time = 3
            elif row_week_day == 4:
                ex_col = 4
                ex_col_time = 5
            elif row_week_day == 5:
                ex_col = 6
                ex_col_time = 7
            else:
                ex_col = 8
                ex_col_time = 9

        for index, row_row in filter_each_date.iterrows():
            row_start_time = row_row['start_time']
            row_end_time = row_row['end_time']
            row_activity = row_row['activity']
            row_project_num = row_row['project_number']
            if row_project_num == None:
                row_project_num = row_activity
            row_time_code = row_row['time_code']
            row_start_convert = datetime.strptime(row_start_time, '%I:%M:%S %p')
            row_end_convert = datetime.strptime(row_end_time, '%I:%M:%S %p')
            row_time_delta = timedelta(minutes=15)
            ex_start_row_time = datetime.strptime(str(time(6,45,00)), '%I:%M:%S')

            for time_unit in range(0,54):
                if row_start_convert == ex_start_row_time + (row_time_delta * time_unit):
                    ex_row = time_unit + 2
                    # print(row_start_convert.time(),row_project_num,row_time_code, ex_row,ex_col)
                else:
                    pass

            for end_time_unit in range(0,54):
                if row_end_convert == ex_start_row_time + (row_time_delta * end_time_unit):
                    ex_row_end = end_time_unit + 1
                else:
                    pass

            ws.cell(row=ex_row_end,column=ex_col).value = '|'
            ws.cell(row=ex_row,column=ex_col).value = row_project_num
            ws.cell(row=ex_row,column=ex_col_time).value = row_time_code

wb.save('timesheet_to_submit.xlsx')


c.close()
conn.close()
